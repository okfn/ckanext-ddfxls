from io import BytesIO

import openpyxl

from ckanext.ddfxls.writers import xlsx_writer

FIELDS = [
    {'id': 'name', 'type': 'text'},
    {'id': 'age', 'type': 'int'},
    {'id': 'city', 'type': 'text'},
]


def _dump(fields, records):
    with xlsx_writer(fields) as writer:
        writer.write_records(records)
        return writer.end_file()


def _rows(data: bytes):
    worksheet = openpyxl.load_workbook(BytesIO(data)).active
    return [list(row) for row in worksheet.iter_rows(values_only=True)]


def test_headers_follow_field_order():
    rows = _rows(_dump(FIELDS, []))
    assert rows == [['name', 'age', 'city']]


def test_values_align_with_headers_regardless_of_record_key_order():
    records = [
        {'name': 'Ana', 'age': 30, 'city': 'Oslo'},
        # Same data, different key order: must land in the same columns
        {'city': 'Bergen', 'name': 'Bo', 'age': 41},
        {'age': 25, 'city': 'Tromso', 'name': 'Cai'},
    ]
    rows = _rows(_dump(FIELDS, records))
    assert rows == [
        ['name', 'age', 'city'],
        ['Ana', 30, 'Oslo'],
        ['Bo', 41, 'Bergen'],
        ['Cai', 25, 'Tromso'],
    ]


def test_missing_keys_leave_cell_empty_without_shifting_columns():
    records = [{'name': 'Dee', 'city': 'Alta'}]
    rows = _rows(_dump(FIELDS, records))
    # openpyxl reads the '' placeholder cell back as None
    assert rows == [
        ['name', 'age', 'city'],
        ['Dee', None, 'Alta'],
    ]


def test_formula_values_are_stored_as_text():
    records = [{'name': '=HYPERLINK("http://evil.example","x")', 'age': '=1+2', 'city': 'Oslo'}]
    worksheet = openpyxl.load_workbook(BytesIO(_dump(FIELDS, records))).active
    name_cell, age_cell, city_cell = worksheet[2]
    # '='-prefixed values must survive byte-identical but never as live formulas
    assert name_cell.value == '=HYPERLINK("http://evil.example","x")'
    assert name_cell.data_type == 's'
    assert age_cell.value == '=1+2'
    assert age_cell.data_type == 's'
    assert city_cell.data_type == 's'


def test_multiple_write_records_calls_append():
    data = None
    with xlsx_writer(FIELDS) as writer:
        writer.write_records([{'name': 'Ana', 'age': 30, 'city': 'Oslo'}])
        writer.write_records([{'name': 'Bo', 'age': 41, 'city': 'Bergen'}])
        data = writer.end_file()
    rows = _rows(data)
    assert rows == [
        ['name', 'age', 'city'],
        ['Ana', 30, 'Oslo'],
        ['Bo', 41, 'Bergen'],
    ]
