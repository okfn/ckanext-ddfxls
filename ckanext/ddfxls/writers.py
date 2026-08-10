from __future__ import annotations

from typing import Any
from contextlib import contextmanager
from io import BytesIO
import openpyxl
from openpyxl.cell.cell import TYPE_STRING
from openpyxl.utils import get_column_letter


@contextmanager
def xlsx_writer(fields: list[dict[str, Any]], bom: bool = False):
    """Context manager for writing XLSX data to file

    :param fields: list of datastore fields
    :param bom: ignored for XLSX format
    """

    output = BytesIO()
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Data"

    # Write headers
    for idx, field in enumerate(fields, 1):
        col_letter = get_column_letter(idx)
        worksheet[f'{col_letter}1'] = field['id']

    yield XLSXWriter(output, workbook, worksheet, fields)


class XLSXWriter(object):
    def __init__(self, output: BytesIO, workbook: Any,
                 worksheet: Any, fields: list[dict[str, Any]]):
        self.output = output
        self.workbook = workbook
        self.worksheet = worksheet
        self.fields = fields
        self.current_row = 2  # Start after header row

    def write_records(self, records: list[dict[str, Any]]) -> bytes:
        """Write records to the XLSX worksheet"""

        for record in records:
            for idx, field in enumerate(self.fields, 1):
                cell = self.worksheet.cell(row=self.current_row, column=idx)
                value = record.get(field['id'], '')
                cell.value = value
                if isinstance(value, str) and value.startswith('='):
                    # openpyxl stores '='-prefixed strings as live formulas;
                    # force text to prevent formula injection
                    cell.data_type = TYPE_STRING
            self.current_row += 1

        return b''  # No incremental output for XLSX

    def end_file(self) -> bytes:
        """Finalize the XLSX file and return the complete file"""
        self.workbook.save(self.output)
        self.output.seek(0)
        return self.output.read()
